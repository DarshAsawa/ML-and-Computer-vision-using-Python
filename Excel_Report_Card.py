import openpyxl

#WRITING TO EXCEL SHEET......
wb=openpyxl.Workbook()
sheet = wb.active


for row in sheet['A1:B3']:
  for cell in row:
    cell.value = None

no_subjects=int(input("Enter the number of subjects:"))
for i in range(2,no_subjects+2):
    c=sheet.cell(row=i,column=1)
    c1=sheet.cell(row=i,column=2)
    subject=input("Enter the name of the Subject: ")
    c.value=subject
    score=input("Enter the marks for %s: " % subject)
    c1.value=score

wb.save("C:\\Users\\dasaw\\Documents\\Python_Github\\test.xlsx")


# READING FROM EXCEL SHEET.....
path = "C:\\Users\\dasaw\\Documents\\Python_Github\\Report_Card.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

print("Subject : Marks")

max_marks=0
mrow=sheet.max_row
total=0

for i in range(2,mrow+1):
    cell_obj1=sheet.cell(row=i,column=1)
    cell_obj=sheet.cell(row=i,column=2)
    print(cell_obj1.value," : ",cell_obj.value)
    total+=100
    max_marks+=cell_obj.value

print("Marks scored out of %d: %d" % (total,max_marks))
print("Percentage scored:")
#cell = sheet.cell(row=1 , column=1)

