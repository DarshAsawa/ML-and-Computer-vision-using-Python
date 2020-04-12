# READING FROM EXCEL SHEET.....

import openpyxl

path = "C:\\Users\\dasaw\\Documents\\Python_Github\\Report_Card.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

c1=sheet["A1"]
c2=sheet["B1"]

print("(%s : %s)\n" % (c1.value,c2.value))

max_marks=0
mrow=sheet.max_row
total=0

for i in range(2,mrow+1):
    cell_obj1=sheet.cell(row=i,column=1)
    cell_obj=sheet.cell(row=i,column=2)
    print(cell_obj1.value," : ",cell_obj.value)
    max_marks+=100
    total+=cell_obj.value

marks_scored=total/max_marks
print("\nMarks scored out of %d: %d" % (max_marks,total))
print("Percentage scored: %d %%" % (marks_scored*100))

